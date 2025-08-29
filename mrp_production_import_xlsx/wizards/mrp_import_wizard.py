# Copyright 2025 bring.out doo Sarajevo
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).

import base64
import io
import logging
from contextlib import contextmanager
from datetime import datetime

from openpyxl import load_workbook

from odoo import api, fields, models, tools, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class MrpProductionImport(models.TransientModel):
    _name = "mrp.production.import"
    _description = "Import Manufacturing Orders from XLS/XLSX"

    file = fields.Binary(string="XLSX file", required=True)
    file_name = fields.Char("File Name")
    product_id = fields.Many2one(
        comodel_name="product.product",
        string="Product to produce",
        required=True,
    )
    product_qty = fields.Float(
        string="Quantity to produce",
        required=True,
        default=1.0,
    )
    product_id = fields.Many2one(
        comodel_name="product.product",
        string="Product to produce",
        required=True,
    )
    product_qty = fields.Float(
        string="Quantity to produce",
        required=True,
        default=1.0,
    )
    auto_confirm = fields.Boolean(
        string="Automatically confirm orders",
        default=False,
        help="If ticked, the wizard will confirm each new MO after creation.",
    )
    log = fields.Text(string="Import log", readonly=True)

    # -----------------------------------------------------------------
    # New confirmation helper fields
    # -----------------------------------------------------------------
    state = fields.Selection(
        [('draft', "Draft"), ("confirm", "Confirm"), ("done", "Done")],
        default="draft",
        string="Status",
        readonly=True,
    )
    row_count = fields.Integer(
        string="Rows in file",
        readonly=True,
        help="Number of non-empty data rows detected in the spreadsheet.",
    )
    progress_current = fields.Integer(
        string="Current Progress",
        readonly=True,
        default=0,
        help="Number of rows processed so far"
    )
    progress_total = fields.Integer(
        string="Total Progress",
        readonly=True,
        default=0,
        help="Total number of rows to process"
    )
    progress_percentage = fields.Float(
        string="Progress %",
        compute="_compute_progress_percentage",
        readonly=True
    )
    is_processing = fields.Boolean(
        string="Is Processing",
        readonly=True,
        default=False,
        help="Indicates if import is currently in progress"
    )
    import_status = fields.Selection([
        ('success', 'Success'),
        ('error', 'Error'),
    ], string="Import Status", readonly=True)

    @api.depends('progress_current', 'progress_total')
    def _compute_progress_percentage(self):
        for record in self:
            if record.progress_total > 0:
                record.progress_percentage = (record.progress_current / record.progress_total) * 100
            else:
                record.progress_percentage = 0.0


    # ---------------------------------------------------------------------
    # Public buttons
    # ---------------------------------------------------------------------

    # Step-1: user clicks “Continue” → we just count the rows
    def action_count_rows(self):
        self.ensure_one()

            # Validate file type first
        if not self.file_name or not self.file_name.lower().endswith(('.xlsx', '.xls')):
            raise UserError(_("Please upload a valid XLS/XLSX file"))


        rows = list(self._read_xlsx())          # iterator -> list so we can len()
        self.row_count = len(rows)
        if not self.row_count:
            raise UserError(_("The file does not contain any data rows."))
        self.state = "confirm"
        return self._action_reload()

    def action_import(self):
        """
        Real import – only reachable after the user acknowledged the amount
        of rows (state = *confirm*).
        """
        self.ensure_one()
        if self.state != "confirm":
            raise UserError(_("Please confirm the import first."))

        # Set processing flag and initialize progress
        self.write({
            'is_processing': True,
            'progress_current': 0,
            'progress_total': self.row_count,
        })
        self.env.cr.commit()  # Commit to make progress visible

        try:
            # Turn the iterator into a *list* so we can iterate twice
            rows = list(self._read_xlsx())

            # --- NEW  : preliminary validation of product codes -------------
            missing_codes = self._check_missing_products(rows)
            if missing_codes:
                self.log = _(
                    "Import aborted — the following product codes do not exist "
                    "in the database:\n%s"
                ) % ", ".join(sorted(missing_codes))
                self.is_processing = False
                self.import_status = 'error'
                self.state = 'done'
                return self._action_show_log()
            # ----------------------------------------------------------------

            mo, warehouse = self._create_mo()
            ok, ko = self._process_rows(rows, mo, warehouse)

            log_lines = [_("1 production order created.")]
            if ko:
                log_lines.append(_("Errors (%s):") % len(ko))
                log_lines.extend(ko)
                self.import_status = 'error'
            else:
                self.import_status = 'success'

            self.log = "\n".join(log_lines)
            self.state = "done"
        finally:
            self.is_processing = False

        return self._action_show_log()

    # -----------------------------------------------------------------
    # Small re-load helper (used by both buttons)
    # -----------------------------------------------------------------
    #def _action_reload(self):
    #    self.ensure_one()
    #    return {
    #        "type": "ir.actions.act_window",
    #        "res_model": "mrp.production.import",
    #        "view_mode": "form",
    #        "res_id": self.id,
    #        "target": "new",
    #    }

    def _action_reload(self):
        """Reload the wizard to refresh the view."""
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
            'context': self.env.context,
        }
    # -----------------------------------------------------------------
    # NEW  : product-code validation *before* row processing
    # -----------------------------------------------------------------
    def _check_missing_products(self, rows):
        """
        Return a *set* of default_codes that are present in the spreadsheet
        but not in the database. The search is done in bulk for efficiency.
        """
        wanted_refs = {str(row.get(_("Product"))).strip() for row in rows if row.get(_("Product"))}
        if not wanted_refs:
            raise UserError(_("Column 'Product' is empty in the file."))

        Product = self.env["product.product"].with_context(active_test=False)
        
        # Perform a single search to find all matching products
        domain = [
            #'|', '|',
            ('default_code', 'in', list(wanted_refs)),
            #('barcode', 'in', list(wanted_refs)),
            #('name', 'in', list(wanted_refs))
        ]
        existing_products = Product.search(domain)

        # Extract the references from the found products
        found_refs = set()
        for product in existing_products:
            if product.default_code in wanted_refs:
                found_refs.add(product.default_code)
            #if product.barcode in wanted_refs:
            #    found_refs.add(product.barcode)
            #if product.name in wanted_refs:
            #    found_refs.add(product.name)

        return wanted_refs - found_refs

    # ---------------------------------------------------------------------
    # I/O helpers
    # ---------------------------------------------------------------------

    def _read_xlsx(self):
        """Return an *iterator* of dicts: one per data row."""
        self.ensure_one()
        try:
            binary = io.BytesIO(base64.b64decode(self.file))
            wb = load_workbook(binary, read_only=True, data_only=True)
        except Exception as exc:
            raise UserError(_("Unable to read the file: %s") % tools.ustr(exc))

        ws = wb.active
        try:
            headers = [str(cell.value).strip() for cell in next(ws.rows)]
        except StopIteration:
            raise UserError(_("The file is empty."))

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(val is None for val in row):
                continue  # skip empty lines
            yield dict(zip(headers, row))

    # ---------------------------------------------------------------------
    # Row processing
    # ---------------------------------------------------------------------

    def _create_mo(self):
        self.ensure_one()
        bom = self.env["mrp.bom"].search([("product_tmpl_id", "=", self.product_id.product_tmpl_id.id)], limit=1)
        warehouse, src_loc, dest_loc = self._get_warehouse_default_locations()
        mo_vals = {
            "product_id": self.product_id.id,
            "product_qty": self.product_qty,
            "product_uom_id": self.product_id.uom_id.id,
            "bom_id": bom.id,
            "location_src_id": src_loc.id,
            "location_dest_id": dest_loc.id,
        }
        mo = self.env["mrp.production"].create(mo_vals)
        return mo, warehouse

    def _process_rows(self, rows, mo, warehouse):
        ok, ko = 0, []
        total_rows = len(rows)

        # Process in batches for better performance
        batch_size = 10  # Update progress every 10 rows

        for line_no, row in enumerate(rows, start=2):
            try:
                with self._with_savepoint():
                    move_vals = self._prepare_move_vals(row, mo, warehouse)
                    self.env["stock.move"].create(move_vals)
                ok += 1
            except Exception as exc:
                _logger.exception("Manufacturing import error on line %s", line_no)
                ko.append(_("Line %(line)s – %(error)s",
                            line=line_no, error=tools.ustr(exc)))

            # Update progress
            current_row = line_no - 1  # Adjust for 0-based indexing
            if current_row % batch_size == 0 or current_row == total_rows:
                self.write({'progress_current': current_row})
                self.env.cr.commit()  # Commit progress update

        return ok, ko

    @api.model
    def get_import_progress(self, wizard_id):
        """Method to be called via RPC to check progress"""
        wizard = self.browse(wizard_id)
        if wizard.exists():
            return {
                'current': wizard.progress_current,
                'total': wizard.progress_total,
                'percentage': wizard.progress_percentage,
                'is_processing': wizard.is_processing,
                'state': wizard.state,
            }
        return {'error': 'Wizard not found'}

    @contextmanager
    def _with_savepoint(self):
        """Rollback the current row only, not the whole wizard."""
        self.env.cr.execute("SAVEPOINT mrp_import")
        try:
            yield
            self.env.cr.execute("RELEASE SAVEPOINT mrp_import")
        except Exception:
            self.env.cr.execute("ROLLBACK TO SAVEPOINT mrp_import")
            raise

    # ---------------------------------------------------------------------
    # Mapping helpers
    # ---------------------------------------------------------------------

    def _prepare_move_vals(self, row, mo, warehouse):
        product = self._get_product(row.get(_("Product")))
        qty = self._float(row.get(_("Quantity"), 0.0))
        if qty <= 0.0:
            raise UserError(_("Quantity must be strictly positive."))

        uom = self._get_uom(row.get(_("UoM"))) or product.uom_id
        return {
            "name": product.name,
            "product_id": product.id,
            "product_uom_qty": qty,
            "product_uom": uom.id,
            "location_id": mo.location_src_id.id,
            "location_dest_id": mo.location_dest_id.id,
            "raw_material_production_id": mo.id,
            "company_id": mo.company_id.id,
            "warehouse_id": warehouse.id,
        }

    

    # ------ elementary converters ----------------------------------------

    @staticmethod
    def _float(value):
        if value is None:
            return 0.0
        try:
            return float(str(value).replace(",", "."))
        except Exception:
            raise UserError(_("Could not convert %s to a number.") % value)

    @staticmethod
    def _date(value):
        if not value:
            return False
        if isinstance(value, datetime):
            return value
        try:
            # Try Odoo's built-in parser first
            return fields.Datetime.from_string(value)
        except Exception:
            # Fall back to ISO pattern
            try:
                return datetime.strptime(value, "%Y-%m-%d")
            except Exception:
                raise UserError(_("Bad date: %s") % value)

    # ------ record lookup helpers ----------------------------------------

    def _get_warehouse_default_locations(self):
        
        picking_type_id = self.env['stock.picking.type'].search(
            [
                ('company_id', '=', self.env.company.id),
                ('code', '=', 'mrp_operation'),
                ('active', '=', True)
            ]
        )
        picking_type_id.ensure_one()
        
        warehouse = picking_type_id.warehouse_id
        #warehouse = self.env['stock.warehouse'].search([('company_id', '=', self.env.company.id)], limit=1)
        if not warehouse:
            return None, None
        rule = warehouse.pbm_route_id.rule_ids and warehouse.pbm_route_id.rule_ids[0] or None
        if not rule:
            return None, None
        #return rule.location_src_id, rule.location_dest_id
        return warehouse, rule.location_dest_id, rule.location_src_id
        
    def _get_product(self, ref):
        if not ref:
            raise UserError(_("Column 'Product' is mandatory."))
        Product = self.env["product.product"].with_context(active_test=False)

        product = Product.search(
            [ 
             #"|", 
             ("default_code", "=", ref), 
             #("barcode", "=", ref)
            ], limit=1
        )
        #if not product:
        #    product = Product.search([("name", "=", ref)], limit=1)
        if not product:
            raise UserError(_("Product not found: %s") % ref)
        return product

    def _get_uom(self, name):
        if not name:
            return None
        Uom = self.env["uom.uom"].with_context(active_test=False)
        uom = Uom.search([("name", "=", name)], limit=1)
        if not uom:
            raise UserError(_("UoM not found: %s") % name)
        return uom

    def _get_company(self, name):
        if not name:
            return None
        Company = self.env["res.company"]
        company = Company.search([("name", "=", name)], limit=1)
        if not company:
            raise UserError(_("Company not found: %s") % name)
        return company

    def _get_location(self, name):
        Location = self.env["stock.location"]
        loc = Location.search([("name", "=", name)], limit=1)
        #if not loc:
        #    raise UserError(_("Location not found: %s") % name)
        return loc

    def _get_user(self, name):
        User = self.env["res.users"]
        user = User.search([("name", "=", name)], limit=1)
        if not user:
            raise UserError(_("User not found: %s") % name)
        return user

    # ---------------------------------------------------------------------
    # UI helpers
    # ---------------------------------------------------------------------

    def _action_show_log(self):
        """Reload the wizard so the log becomes visible."""
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "res_model": "mrp.production.import",
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
        }