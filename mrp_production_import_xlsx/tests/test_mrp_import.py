# Copyright 2024 Your Company
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).

import base64
import io
from datetime import date

from openpyxl import Workbook

# AI:  Yes, from odoo.tests.common import TransactionCase 
# is the correct way to import TransactionCase in Odoo 16. 
# It's the standard for creating unit tests that interact with the database.

from odoo.tests.common import TransactionCase


class TestMrpProductionImport(TransactionCase):
    """End-to-end test: build an XLSX in memory, run the wizard, check result."""

    def setUp(self):
        super().setUp()
        # -----------------------------------------------------------------
        # Minimal master-data needed by the import
        # -----------------------------------------------------------------
        self.uom_unit = self.env.ref("uom.product_uom_unit")
        self.company = self.env.company
        self.location_src = self.env.ref("stock.stock_location_stock")
        self.location_dest = self.env.ref("stock.stock_location_finished_products")

        # Two products that will be referenced by `default_code`
        self.prod_cable = self.env["product.product"].create(
            {
                "name": "Cable",
                "default_code": "P-CABLE",
                "type": "product",
                "uom_id": self.uom_unit.id,
                "uom_po_id": self.uom_unit.id,
            }
        )
        self.prod_board = self.env["product.product"].create(
            {
                "name": "Main board",
                "default_code": "P-BOARD",
                "type": "product",
                "uom_id": self.uom_unit.id,
                "uom_po_id": self.uom_unit.id,
            }
        )

    # ---------------------------------------------------------------------
    # Helper: build an XLSX identical to the README example
    # ---------------------------------------------------------------------
    def _build_sample_xlsx_b64(self):
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "Product",
                "Quantity",
                "UoM",
                "Planned Start",
                "Reference",
                "Company",
                "Source Location",
                "Destination Location",
                "Responsible",
                "Priority",
            ]
        )
        ws.append(
            [
                "P-CABLE",
                10,
                "Unit",
                date(2024, 6, 15),
                "MO-XLS-01",
                self.company.name,
                self.location_src.name,
                self.location_dest.name,
                "Administrator",
                1,
            ]
        )
        ws.append(
            [
                "P-BOARD",
                5,
                "Unit",
                date(2024, 7, 1),
                "MO-XLS-02",
                self.company.name,
                None,
                None,
                None,
                None,
            ]
        )

        buffer = io.BytesIO()
        wb.save(buffer)
        return base64.b64encode(buffer.getvalue())

    # ---------------------------------------------------------------------
    # The actual unit-test
    # ---------------------------------------------------------------------
    def test_import_manufacturing_orders(self):
        file_b64 = self._build_sample_xlsx_b64()

        wizard = self.env["mrp.production.import"].create(
            {
                "file": file_b64,
                "file_name": "sample_mrp_import.xlsx",
                "auto_confirm": True,  # we also exercise the confirmation path
            }
        )

        wizard.action_import()

        # Two rows ⇒ two Manufacturing Orders
        mos = self.env["mrp.production"].search(
            [("origin", "in", ("MO-XLS-01", "MO-XLS-02"))]
        )
        self.assertEqual(
            len(mos), 2, "The wizard should have created exactly two MOs."
        )

        # Quick sanity check on the first MO
        mo1 = mos.filtered(lambda m: m.origin == "MO-XLS-01")
        self.assertTrue(mo1, "MO-XLS-01 was not created")
        self.assertEqual(mo1.product_id, self.prod_cable)
        self.assertEqual(mo1.product_qty, 10)
        # auto_confirm=True ⇒ state should at least have left 'draft'
        self.assertNotEqual(mo1.state, "draft", "MO was not confirmed")

        # The wizard must store a non-empty log
        self.assertTrue(wizard.log, "Import wizard did not record a log.")

    def test_import_aborts_on_unknown_product(self):
        # Build a file that references one *unknown* code
        file_b64 = self._build_sample_xlsx_b64()
        # tamper with first row
        import base64, io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(base64.b64decode(file_b64)))
        ws = wb.active
        ws["A2"].value = "BAD-SKU-1"          # <- unknown code
        buf = io.BytesIO()
        wb.save(buf)
        file_b64 = base64.b64encode(buf.getvalue())

        wiz = self.env["mrp.production.import"].create(
            {"file": file_b64, "file_name": "bad.xlsx"}
        )
        wiz.action_import()

        self.assertIn("BAD-SKU-1", wiz.log)
        mos = self.env["mrp.production"].search([("origin", "=", "MO-XLS-01")])
        self.assertFalse(mos, "No MO should be created when codes are missing")        